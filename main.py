import os
import re
from threading import Thread

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.utils import platform
from kivy.core.window import Window
from kivy.clock import mainthread

from openpyxl import load_workbook

try:
    from plyer import vibrator
except Exception:
    vibrator = None


def to_bond_number(value):
    """Bond as an integer value, or None. Integer matching makes leading
    zeros safe: Excel stores 033313 as 33313, the winning list shows
    '033313' -> both become 33313 and match. Also strips commas / '.0'."""
    if value is None:
        return None
    digits = re.sub(r'\D', '', str(value).strip())
    if not digits:
        return None
    n = int(digits)
    return n if n > 0 else None


def as_bond_text(n):
    """Display padded to 6 digits, the way bonds are printed."""
    return str(n).zfill(6)


class BondMatcherApp(App):
    def build(self):
        self.title = "Bond Matcher"
        self.txt_files = []      # winning lists
        self.excel_files = []    # your purchased bonds

        self.layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self.btn_txt = Button(text="BROWSE WINNING LISTS (.txt)", size_hint_y=None,
                              height=60, background_color=(0.2, 0.6, 1, 1))
        self.btn_txt.bind(on_release=lambda x: self.browse_files(mode='txt'))

        self.btn_excel = Button(text="BROWSE YOUR BONDS (.xlsx)", size_hint_y=None,
                                height=60, background_color=(0.2, 0.6, 1, 1))
        self.btn_excel.bind(on_release=lambda x: self.browse_files(mode='excel'))

        self.status_label = Label(text="Files Selected:  Bonds: 0  |  Winning Lists: 0",
                                  size_hint_y=None, height=40)

        self.scroll = ScrollView()
        self.result_label = Label(text="Browse your bond files and winning lists, then press START SCAN.",
                                  size_hint_y=None, halign='left', valign='top',
                                  padding=(10, 10), markup=True)
        self.result_label.bind(
            width=lambda *a: setattr(self.result_label, 'text_size',
                                     (self.result_label.width, None)))
        self.result_label.bind(
            texture_size=lambda *a: setattr(self.result_label, 'height',
                                            self.result_label.texture_size[1]))
        self.scroll.add_widget(self.result_label)

        self.actions = BoxLayout(orientation='horizontal', size_hint_y=None,
                                 height=80, spacing=10)
        self.start_btn = Button(text="START SCAN", background_color=(0, 0.8, 0, 1),
                                font_size='18sp', bold=True)
        self.start_btn.bind(on_press=self.start_scan)
        self.clear_btn = Button(text="CLEAR", background_color=(0.5, 0.5, 0.5, 1),
                                font_size='18sp', bold=True)
        self.clear_btn.bind(on_press=self.clear_all)
        for b in (self.start_btn, self.clear_btn):
            self.actions.add_widget(b)
        if platform != 'android':
            self.exit_btn = Button(text="EXIT", background_color=(0.8, 0, 0, 1),
                                   font_size='18sp', bold=True)
            self.exit_btn.bind(on_press=self.exit_app)
            self.actions.add_widget(self.exit_btn)

        for w in (self.btn_txt, self.btn_excel, self.status_label, self.scroll, self.actions):
            self.layout.add_widget(w)

        if platform == 'android':
            self._android_setup()
        return self.layout

    # ---------------------------------------------------------- Android setup
    def _android_setup(self):
        """Ask for storage access so we can read files the user picks."""
        try:
            from android.permissions import request_permissions, Permission
            request_permissions([Permission.READ_EXTERNAL_STORAGE,
                                 Permission.WRITE_EXTERNAL_STORAGE])
        except Exception:
            pass
        # Android 11+ needs "All files access" to read arbitrary files (e.g. Download).
        try:
            from jnius import autoclass
            version = autoclass('android.os.Build$VERSION')
            if version.SDK_INT >= 30:
                Environment = autoclass('android.os.Environment')
                if not Environment.isExternalStorageManager():
                    Intent = autoclass('android.content.Intent')
                    Settings = autoclass('android.provider.Settings')
                    Uri = autoclass('android.net.Uri')
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                    act = PythonActivity.mActivity
                    intent = Intent(Settings.ACTION_MANAGE_APP_ALL_FILES_ACCESS_PERMISSION)
                    intent.setData(Uri.parse("package:" + act.getPackageName()))
                    act.startActivity(intent)
        except Exception:
            pass

    # ---------------------------------------------------------------- browsing
    def browse_files(self, mode):
        if platform == 'android':
            self._android_browse(mode)
        elif platform in ('win', 'macosx', 'linux'):
            self._desktop_browse(mode)
        else:
            self._kivy_picker(mode)

    def _android_browse(self, mode):
        """Native Android file picker (multi-select)."""
        try:
            from plyer import filechooser
            filechooser.open_file(
                multiple=True,
                on_selection=lambda sel: self._android_selected(mode, sel))
        except Exception:
            self._kivy_picker(mode)   # fall back to in-app browser

    @mainthread
    def _android_selected(self, mode, selection):
        if selection:
            self._set_selection(mode, list(selection))

    def _desktop_browse(self, mode):
        """Native OS 'Open File' dialog on PC (multi-select)."""
        try:
            import tkinter
            from tkinter import filedialog
            root = tkinter.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            if mode == 'txt':
                ftypes, title = [("Text files", "*.txt"), ("All files", "*.*")], "Select Winning List file(s)"
            else:
                ftypes, title = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")], "Select Your Bond file(s)"
            selection = filedialog.askopenfilenames(title=title, filetypes=ftypes)
            root.destroy()
            if selection:
                self._set_selection(mode, list(selection))
        except Exception:
            self._kivy_picker(mode)

    def _kivy_picker(self, mode):
        content = BoxLayout(orientation='vertical')
        filters = ['*.txt'] if mode == 'txt' else ['*.xlsx', '*.xls']
        chooser = FileChooserListView(multiselect=True, filters=filters)
        if platform == 'android':
            chooser.path = "/storage/emulated/0/Download"
        bar = BoxLayout(size_hint_y=None, height=60)
        ok, cancel = Button(text="CONFIRM"), Button(text="CANCEL")
        bar.add_widget(ok); bar.add_widget(cancel)
        content.add_widget(chooser); content.add_widget(bar)
        popup = Popup(title=f"Select {mode.upper()} file(s)", content=content, size_hint=(0.95, 0.95))
        ok.bind(on_release=lambda *_: (self._set_selection(mode, list(chooser.selection)), popup.dismiss()))
        cancel.bind(on_release=popup.dismiss)
        popup.open()

    def _set_selection(self, mode, files):
        if mode == 'txt':
            self.txt_files = files
        else:
            self.excel_files = files
        self.status_label.text = (f"Files Selected:  Bonds: {len(self.excel_files)}  |  "
                                  f"Winning Lists: {len(self.txt_files)}")

    def clear_all(self, *_):
        self.txt_files, self.excel_files = [], []
        self.status_label.text = "Files Selected:  Bonds: 0  |  Winning Lists: 0"
        self.set_result("Cleared. Browse files again to start.")

    def exit_app(self, *_):
        App.get_running_app().stop()
        Window.close()

    # ------------------------------------------------------------------- scan
    def start_scan(self, *_):
        if not self.excel_files or not self.txt_files:
            self.set_result("[color=ff0000]Please select at least one BOND file and one WINNING LIST.[/color]")
            return
        self.start_btn.disabled = True
        self.set_result("Comparing files... please wait.")
        Thread(target=self.run_search, daemon=True).start()

    @mainthread
    def set_result(self, text):
        self.result_label.text = text

    @mainthread
    def finish_scan(self):
        self.start_btn.disabled = False

    def _read_bonds(self, path):
        """Read every numeric cell from an .xlsx using openpyxl (no pandas)."""
        numbers = []
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        n = to_bond_number(cell)
                        if n is not None:
                            numbers.append(n)
        finally:
            wb.close()
        return numbers

    def run_search(self):
        errors = []
        try:
            bond_source = {}
            for path in self.excel_files:
                name = os.path.basename(path)
                if not os.path.exists(path):
                    errors.append(f"{name}: file not readable (move it to Download and grant 'All files access').")
                    continue
                try:
                    for n in self._read_bonds(path):
                        bond_source[n] = name
                except Exception as e:
                    errors.append(f"{name}: {e}")

            if not bond_source:
                msg = "No bond numbers found in the selected Excel file(s)."
                if errors:
                    msg += "\n\n[color=ff0000]" + "\n".join(errors) + "[/color]"
                self.set_result(msg)
                return

            matches = {}
            for path in self.txt_files:
                name = os.path.basename(path)
                if not os.path.exists(path):
                    errors.append(f"{name}: file not readable (move it to Download and grant 'All files access').")
                    continue
                try:
                    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                        for line in f:
                            for token in re.findall(r'\d+', line):
                                n = int(token)
                                if n in bond_source and n not in matches:
                                    matches[n] = (name, bond_source[n])
                except Exception as e:
                    errors.append(f"{name}: {e}")

            if matches:
                lines = [f"[b][color=00aa00]CONGRATULATIONS! {len(matches)} WINNING BOND(S) FOUND[/color][/b]\n"]
                for n in sorted(matches):
                    win_file, bond_file = matches[n]
                    lines.append(f"[b][color=ff0000]WINNER: {as_bond_text(n)}[/color][/b]\n"
                                 f"- Your bond from: {bond_file}\n"
                                 f"- Found in list: {win_file}\n---")
                self.set_result("\n".join(lines))
                if platform == 'android' and vibrator is not None:
                    try:
                        vibrator.vibrate(time=1)
                    except Exception:
                        pass
            else:
                msg = (f"[b]No number is matched.[/b]\n\n"
                       f"Checked {len(bond_source)} of your bonds against the winning list(s).")
                if errors:
                    msg += "\n\n[color=ff0000]" + "\n".join(errors) + "[/color]"
                self.set_result(msg)
        except Exception as e:
            self.set_result(f"[color=ff0000]System Error: {e}[/color]")
        finally:
            self.finish_scan()


if __name__ == '__main__':
    BondMatcherApp().run()
